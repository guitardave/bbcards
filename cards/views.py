import datetime

from django.conf import settings
from django.http import HttpResponse
from django.views import View
from xhtml2pdf import pisa
import os
from typing import Any, List, Dict

import openpyxl
from django.db.models.functions import Cast
from django.utils import timezone
from openpyxl.workbook import Workbook
import boto3
from django.contrib.auth.decorators import login_required
from django.contrib.postgres.search import SearchVector, SearchQuery
from django.db.models import Q, QuerySet, IntegerField
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView
from django.contrib import messages

from players.models import Player
from .forms import CardSetForm, CardUpdateForm, CardCreateForm, SearchForm
from .models import Card, CardSet, CardListExport
from decorators.my_decorators import error_handling


def card_set_list_fn(request, n_count: int):
    if n_count > 0:
        cards = CardSet.all_sets.all().order_by('-date_entered')[:n_count]
    else:
        cards = CardSet.all_sets.all().order_by('year', 'card_set_name')

    c_data = CardListData(cards)
    card_sets = c_data.card_list_count(True)
    set_count, rs, n_pages = card_list_pagination(request, card_sets, settings.DEFAULT_LIMIT)

    return {
        'rs': rs,
        'title': 'Card Sets',
        'form': CardSetForm,
        'card_title': 'Add Card Set',
        'set_count': set_count,
        'n_pages': n_pages,
        'loaded': timezone.now()
    }


@login_required(login_url='/users/')
@error_handling
def card_set_create_async(request):
    c_message = ''
    if request.method == 'POST':
        form = CardSetForm(request.POST)
        if form.is_valid():
            set_year = form.cleaned_data['year']
            set_name = form.cleaned_data['card_set_name']
            sport = form.cleaned_data['sport']
            full_set_name = f'{set_year} {set_name} {sport}'
            check = CardSet.objects.filter(card_set_name=set_name, year=set_year, sport=sport)
            if not check.exists():
                form.save()
                c_message = f'<i class="fa fa-check"></i> {full_set_name} has been added'
            else:
                c_message = f'<i class="fa fa-remove"></i> {full_set_name} already exists'
        else:
            c_message = f'<i class="fa fa-remove">{form.errors}</i>'
    cards = CardSet.all_sets.all().order_by('-id')[:CardSet.LIMIT]
    c_data = CardListData(cards)
    cards = c_data.card_list_count(True)
    set_count, rs, n_pages = card_list_pagination(request, cards, CardSet.LIMIT)
    last_id = CardSet.objects.last().id
    context = {
        'rs': rs,
        'new_id': last_id,
        'title': 'Card Sets',
        'set_count': set_count,
        'n_pages': n_pages,
        'c_message': c_message
    }
    return render(request, 'cards/cardset-list-card-partial.html', context)


@error_handling
def card_set_list(request, n_count: int = 0, inc_zero: bool = False):
    if request.method == 'POST':
        form = CardSetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Card set has been created')
        return redirect('cards:cardsets', n_count=n_count if n_count > 0 else CardSet.LIMIT)
    context = card_set_list_fn(request, n_count)
    return render(request, 'cards/cardset-list.html', context)


@login_required(login_url="/users/")
@error_handling
def card_set_update_async(request, slug: str):
    obj = CardSet.objects.get(slug=slug)
    if request.method == 'POST':
        form = CardSetForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            t_message = '<i class="fa fa-check"></i>'
        else:
            t_message = '<i class="fa fa-remove"></i> Error'
        context = {
            'card':
                {
                    'card': obj,
                    'count': Card.objects.filter(card_set_id=obj.id).count()
                },
            'success': True,
            't_message': t_message
        }
        return render(request, 'cards/cardset-list-tr-partial.html', context)
    context = {
        'form': CardSetForm(instance=obj),
        'obj': obj,
        'card_title': 'Update Card Set',
        'loaded': timezone.now()
    }
    return render(request, 'cards/cardset-form.html', context)


@login_required(login_url='/users/')
@csrf_exempt
@error_handling
def card_set_delete_async(request, slug: str):
    obj = CardSet.objects.filter(slug=slug)
    t_message = ''
    if obj.exists():
        obj = obj[0]
        if not Card.objects.filter(card_set_id=obj).exists():
            obj.delete()
            t_message = 'Set Deleted'
        else:
            t_message = 'Set not empty'
    context = {'t_message': t_message, **card_set_list_fn(request, CardSet.LIMIT)}
    return render(request, 'cards/cardset-list-card-partial.html', context)


@login_required(login_url='/users/')
@error_handling
def card_set_form_refresh(request):
    context = {'card_title': 'Add Card Set', 'form': CardSetForm, 'loaded': datetime.datetime.now()}
    return render(request, 'cards/cardset-form.html', context)


class CardListData:

    def __init__(self, qs: QuerySet):
        self.qs = qs

    def card_list_context(self, request) -> dict[str, str]:
        d_card_list = self.card_list_dict()
        request.session['rs'] = d_card_list
        return dict(
            rs=self.qs,
            loaded=timezone.now(),
            rs_dict=request.session['rs'] if 'rs' in request.session else []
        )

    def card_list_dict(self) -> list[dict[str, Any]]:
        return [
            {
                'player_name': c.player_id.player_fname + ' ' + c.player_id.player_lname,
                'year': c.card_set_id.year,
                'set_name': c.card_set_id.card_set_name,
                'info': c.card_subset,
                'num': c.card_num
            } for c in self.qs
        ]

    def card_list_count(self, inc_zero: bool = False) -> list[dict]:
        c_list = []
        for card_set in self.qs:
            c_count = Card.objects.filter(card_set_id_id=card_set.id).count()
            if c_count > 0 or inc_zero:
                c_list.append(dict(card=card_set, count=c_count))
        return c_list


@error_handling
def card_list_pagination(request, cards: QuerySet | list[dict], n_count: int = None):
    p = Paginator(cards, n_count if n_count else 100)
    page_number = request.GET['page'] if 'page' in request.GET else 1
    try:
        rs = p.get_page(page_number)
    except PageNotAnInteger:
        rs = p.page(1)
    except EmptyPage:
        rs = p.page(p.num_pages)
    return len(cards), rs, p.num_pages


# @login_required(login_url='/users/')
@error_handling
def card_list_last_n(request):
    cards = Card.last_50.all()
    card_count, rs, n_pages = card_list_pagination(request, cards)
    return render(
        request,
        'cards/card-list.html',
        {
            'title': f'Last {len(cards)} Cards',
            'rs': rs,
            'n_pages': n_pages,
            'card_count': card_count,
            'form': CardCreateForm,
            'card_title': 'Add Card'
        }
    )


# @login_required(login_url='/users/')
@error_handling
def card_list_by_player(request, slug: str):
    obj = Player.objects.get(slug=slug)
    cards = Card.objects.filter(
        player_id__slug=slug
    ).order_by('card_set_id__year', 'card_set_id__slug')
    card_count, rs, n_pages = card_list_pagination(request, cards)
    return render(
        request,
        'cards/card-list.html',
        {
            'title': f'{obj.player_fname} {obj.player_lname}',
            'rs': rs,
            'n_pages': n_pages,
            'card_count': card_count,
            'player': obj,
            'form': CardCreateForm(**{'player': obj.slug}),
            'card_title': 'Add Card - Player'
        }
    )


# @login_required(login_url='/users/')
@error_handling
def card_list_by_set(request, slug: str):
    obj = CardSet.objects.get(slug=slug)
    cards = Card.objects.filter(
        card_set_id__slug=slug
    ).order_by(
        'card_num'
    )
    card_count, rs, n_pages = card_list_pagination(request, cards)
    return render(
        request,
        'cards/card-list.html',
        {
            'title': f'{obj.__str__()}',
            'rs': rs,
            'n_pages': n_pages,
            'card_count': card_count,
            'card_set': obj,
            'form': CardCreateForm(**{'set': obj.slug}),
            'card_title': 'Add Card - Cat Set'
        }
    )


# @login_required(login_url='/users/')
@error_handling
def card_list_all(request, sort_by: int = None):
    if not sort_by:
        sort_by = 0
    if sort_by == 1:
        cards = Card.objects.all().order_by(
            'card_set_id__year',
            'card_set_id__card_set_name',
            'card_num'
        )
    else:
        cards = Card.objects.all().order_by(
            'player_id__player_lname',
            'card_set_id__year',
            'card_set_id__card_set_name'
        )
    card_count, rs, n_pages = card_list_pagination(request, cards, 200)
    return render(
        request,
        'cards/card-list.html',
        {
            'title': 'All Cards List',
            'rs': rs,
            'n_pages': n_pages,
            'card_count': card_count,
            'form': CardCreateForm,
            'card_title': 'Add Card'
        }
    )


@error_handling
def load_cards_async(request):
    card_count, rs, n_pages = card_list_pagination(request)
    context = {'rs': rs, 'loaded': timezone.now(), 'n_pages': n_pages, 'card_count': card_count}
    return render(request, 'cards/card-list-table-partial.html', context)


@login_required(login_url="/users/")
@error_handling
def card_update_async(request, slug: str):
    obj = Card.objects.get(slug=slug)
    if request.method == 'POST':
        success = False

        form = CardUpdateForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            success = True
            t_message = '<i class="fa fa-check"></i>'
        else:
            t_message = '<i class="fa fa-remove"></i> Error'
        return render(
            request,
            'cards/card-list-tr-partial.html',
            {'card': Card.objects.get(id=obj.id), 'success': success, 't_message': t_message}
        )
    context = {
        'form': CardUpdateForm(instance=obj),
        'obj': obj,
        'card_title': 'Update Card',
        'loaded': timezone.now()
    }
    return render(request, 'cards/card-form.html', context)


@login_required(login_url='/users/')
@csrf_exempt
@error_handling
def card_delete_async(request, slug: str):
    c_message, cards, player = None, None, None
    obj = Card.objects.filter(slug=slug)
    if obj.exists():
        cards = Card.objects.filter(player_id_id=obj[0].player_id_id)
        player = Player.objects.get(id=obj[0].player_id_id)
        obj.delete()
        c_message = 'Item deleted successfully'
    context = {
        'rs': cards if cards else Card.last_50.all(),
        'c_message': c_message,
        'title': f'Last {len(cards)} Cards' if not player else f'{player.player_fname} {player.player_lname}'
    }
    return render(request, 'cards/card-list-card-partial.html', context)


class TypeSlugs:
    PLAYER = 'player'
    CARD_SET = 'card_set'


@login_required(login_url='/users/')
@error_handling
def card_create_async(request, card_type: str = None, type_slug: str = None):
    player_id, new_id = None, {'id': None}

    form = CardCreateForm(request.POST, request.FILES)
    if form.is_valid():
        form.save()
        new_id = Card.objects.last()
        player_id = Player.objects.get(id=new_id.player_id_id)

    if card_type and type_slug:
        if card_type == TypeSlugs.PLAYER:
            obj = Player.objects.get(slug=type_slug)
            cards = Card.objects.filter(player_id__slug=type_slug).order_by('-id')
            title = f'{obj.player_fname} {obj.player_lname}'
        else:
            obj = CardSet.objects.get(slug=type_slug)
            cards = Card.objects.filter(card_set_id__slug=type_slug).order_by('-id')
            title = f'{str(obj.year)} {obj.card_set_name}'
    elif player_id:
        obj = Player.objects.get(id=player_id.id)
        cards = Card.objects.filter(player_id_id=player_id.id).order_by('-id')
        title = f'{obj.player_fname} {obj.player_lname}'
    else:
        cards = Card.last_50.all()
        title = f'Last {len(cards)} Cards'
    card_count, rs, n_pages = card_list_pagination(request, cards, 150)
    context = {
        'title': title,
        'new_id': new_id.id,
        'card_count': card_count,
        'rs': rs,
        'n_pages': n_pages
    }
    return render(request, 'cards/card-list-card-partial.html', context)


@login_required(login_url='/users/')
@error_handling
def card_create_form_async(request, card_type: str = None, type_slug: str = None):
    form = CardCreateForm
    context = {'card_title': 'Add Card'}
    if card_type and type_slug:
        if card_type == 'player':
            kwargs = {'player': type_slug}
            context['card_title'] = 'Add Card - Player'
        else:
            kwargs = {'set': type_slug}
            context['card_title'] = 'Add Card - Set'
        form = CardCreateForm(**kwargs)
    context['form'] = form
    return render(request, 'cards/card-form.html', context)


@login_required(login_url='/users/')
@error_handling
def card_form_refresh(request):
    context = {'card_title': 'Add New Card', 'form': CardCreateForm, 'loaded': datetime.datetime.now()}
    return render(request, 'cards/card-form.html', context)


@login_required(login_url='/users/')
@error_handling
def card_image(request, slug: str):
    obj = Card.objects.get(slug=slug)
    card_string = f'{obj.card_set_id.year} {obj.card_set_id.card_set_name} {obj.card_subset} {obj.card_num}'
    context = {'title': obj.card_image, 'object': obj, 'card_string': card_string}
    return render(request, 'cards/card-image.html', context)


def card_search_pagination(request, search: str):
    pass


class CardSearch(View):
    template_name = 'cards/card-list-card-partial.html'
    model = Card
    redirect_to = 'cards:card-list-50'

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    def search_query(self, qs: str) -> QuerySet:
        return self.model.objects.search_query(qs).order_by(
            'card_set_id__year',
            'card_set_id__card_set_name',
            'player_id__player_lname'
        )

    def post(self, request, *args, **kwargs):
        form = SearchForm(request.POST)
        if form.is_valid():
            search = form.cleaned_data['search'] if 'search' in form.cleaned_data else self.get(request, *args, **kwargs)
        else:
            search = ''
            messages.warning(request, f'{form.errors}')
        cards = self.search_query(search)
        request.session['rs'] = CardListData(cards).card_list_dict()

        card_list_ctx = CardListData(cards).card_list_context(request)
        card_count, rs, n_pages = card_list_pagination(request, cards, 200)
        context = {
            'title': f'Search "{search}"',
            'form': CardCreateForm,
            'search': search,
            'card_title': 'Add New Card',
            'card_count': card_count,
            'rs': rs,
            'n_pages': n_pages,
            'loaded': card_list_ctx['loaded'],
            'rs_dict': card_list_ctx['rs_dict']
        }
        return render(request, self.template_name, context)

    def get(self, request, *args, **kwargs):
        return redirect(self.redirect_to)


class ExportScriptsExcel:
    DEST_DIR = 'static/bbcards/export_data/'

    def __init__(self, rs: list[dict[str, str]], file_name: str) -> None:
        self.file_name = file_name
        self.rs = rs

    def create_file(self) -> Workbook:
        wb = openpyxl.Workbook()
        wb.save(self.file_name)
        return wb

    def populate_excel_file(self) -> Workbook:
        rs = self.rs
        self.create_file()
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        for i, s in enumerate(rs):
            ws.cell(i + 1, 1).value = s['player_name']
            ws.cell(i + 1, 2).value = s['year']
            ws.cell(i + 1, 3).value = s['set_name']
            ws.cell(i + 1, 4).value = s['info']
            ws.cell(i + 1, 5).value = s['num']
        wb.save(self.file_name)
        return wb

    def upload_excel(self) -> Any:
        wb = self.populate_excel_file()
        client = boto3.Session(os.getenv('AWS_ACCESS_KEY_ID'), os.getenv('AWS_SECRET_ACCESS_KEY')).resource('s3')
        bucket = os.environ.get('AWS_STORAGE_BUCKET_NAME')
        result = client.Bucket(bucket).upload_file(self.file_name, self.DEST_DIR + self.file_name)

        os.remove(self.file_name)

        print(result)
        return result


def export_list_save(request, file_name: str):
    export = CardListExport(user_id=request.user.id, file_name=file_name)
    export.save()


def card_list_export(rs: list[dict] | QuerySet) -> str:
    f_date = datetime.datetime.strftime(timezone.now(), '%m_%d_%Y__%H_%M_%S')
    file_name = f"bbcards_card_list_export_{str(f_date)}.xlsx"
    if len(rs) > 0:
        xl = ExportScriptsExcel(rs, file_name)
        xl.upload_excel()
    return file_name


def card_list_export_pdf(text: str = None) -> str:
    s = CardSearch()
    rs = s.search_query(text)
    html = '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">'
    html += '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
    html += '<style>body, h1, p, th, td {font-family: "Lato", "Calibri", Arial, sans-serif;}</style>'

    html += f'</head><body><div class="container"><h1>Search: {text}</h1><table>'
    for r in rs:
        html += '<tr>'
        html += f'<td>{r.player_id.player_fname} {r.player_id.player_lname}</td>'
        html += f'<td>{r.card_set_id.year}</td>'
        html += f'<td>{r.card_set_id.card_set_name}</td>'
        html += f'<td>{r.card_subset}</td>'
        html += f'<td>{r.card_num}</td>'
        html += '</tr>'
    html += f'</table><p>{len(rs)} records</p>'
    html += f'<hr>Date: {datetime.datetime.strftime(timezone.now(), "%m/%d/%Y %H:%M")}'
    html += '</div></body></html>'
    return html


def html_to_pdf():
    pdf_path = f'bbcards_export__{datetime.datetime.strftime(timezone.now(), "%m%d%Y_%H%M")}.pdf'
    html_content = card_list_export_pdf(None)
    with open(pdf_path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)
    return not pisa_status.err


@login_required(login_url='/users/')
@csrf_exempt
@error_handling
def card_list_export_vw_pdf(request, q: str = None):
    search = ''
    if request.method == 'POST':
        form = SearchForm(request.POST or None)
        if form.is_valid():
            search = form.cleaned_data['search']
        else:
            search = '<span class="fade-out">Invalid input</span>'
    else:
        search = 'Bo Jackson'
    return render(
        request,
        'cards/card-list-export-msg-partial.html',
        {'xport_result': card_list_export_pdf(search), 'pdf': html_to_pdf()}
    )


@login_required(login_url='/users/')
@csrf_exempt
@error_handling
def card_list_export_vw(request):
    rs = card_list_export(request.session['rs']) if 'rs' in request.session else ''
    if len(rs) > 0:
        export_list_save(request, rs)
    context = {
        'xport_result': rs,
        'xport_url': f'https://jojodave.s3.amazonaws.com/static/bbcards/export_data/{rs}'
    }
    return render(request, 'cards/card-list-export-msg-partial.html', context)
